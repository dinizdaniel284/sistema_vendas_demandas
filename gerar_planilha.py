import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.styles.numbers import FORMAT_CURRENCY_BRL_SIMPLE

def gerar_planilha():
    # Dados simulados (público geral: supermercados, lojas, TI etc.)
    dados = {
        'Produto': ['Notebook', 'Arroz 5kg', 'Mouse Sem Fio', 'HD Externo 1TB', 'Caderno 96 folhas'],
        'Categoria': ['Informática', 'Alimentos', 'Acessórios', 'Armazenamento', 'Papelaria'],
        'Quantidade': [3, 20, 15, 5, 50],
        'Preço Unitário': [4300.00, 22.50, 75.90, 310.00, 9.90]
    }

    df = pd.DataFrame(dados)

    # Calcular total de cada item
    df['Total Item'] = df['Quantidade'] * df['Preço Unitário']

    # Criar pasta se não existir
    pasta_dados = 'dados'
    os.makedirs(pasta_dados, exist_ok=True)

    # Nome do arquivo com data
    data_hoje = datetime.now().strftime('%Y-%m-%d')
    caminho_arquivo = os.path.join(pasta_dados, f'planilha_vendas_{data_hoje}.xlsx')

    # Salvar planilha Excel
    df.to_excel(caminho_arquivo, index=False)

    # === APLICAR FORMATAÇÃO NO EXCEL COM openpyxl ===
    wb = load_workbook(caminho_arquivo)
    ws = wb.active

    # Estilos
    fonte_negrito = Font(bold=True)
    alinhamento_centralizado = Alignment(horizontal='center')
    borda_fina = Border(left=Side(style='thin'), right=Side(style='thin'),
                        top=Side(style='thin'), bottom=Side(style='thin'))
    preenchimento = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    # Aplicar estilo no cabeçalho
    for cell in ws[1]:
        cell.font = fonte_negrito
        cell.alignment = alinhamento_centralizado
        cell.fill = preenchimento
        cell.border = borda_fina

    # Formatar colunas de valores como moeda
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=5):
        for cell in row:
            cell.number_format = FORMAT_CURRENCY_BRL_SIMPLE
            cell.border = borda_fina

    # Ajustar largura das colunas automaticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # A, B, C...
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

    # Salvar a planilha final formatada
    wb.save(caminho_arquivo)

    print(f"✅ Planilha personalizada gerada com sucesso em: {caminho_arquivo}")

# Executa
if __name__ == "__main__":
    gerar_planilha()
